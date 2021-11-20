import openpyxl
from mailmerge import MailMerge
from datetime import datetime
from re import split, findall
from docx2pdf import convert

end = '\033[0m'
underline = '\033[4m'


wb = openpyxl.load_workbook('workfile.xlsx',data_only=True)
sheet = wb["Sheet3"]

max_col = sheet.max_row
sheet.delete_rows(sheet.min_row, 1)

template = "invTemp.docx"


def customer_name(Name, l):
    # print(i)
    Product = ""
    Quantity = ""
    Priceperkg = ""
    Subtotal = ""
    TD = 0
    Total = 0
    unit = ""
    for x in range(l, max_col + 1):
        # print(name,str(sheet.cell(row=x, column=1).value))
        if Name == str(sheet.cell(row=x, column=1).value):
            # print(f'This is x={x}')
            # if str(sheet.cell(row=x, column=4).value)=="" \
            #                                            "":
            #     unit=" ltr\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Milk":
            #     unit = " ltr\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Chuijhaal-Half KG":
            #     unit = " pcs\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Chuijhaal-One KG":
            #     unit = " pcs\n"
            #
            # elif str(sheet.cell(row=x, column=4).value)=="Gawa Ghee-400gm":
            #     unit=" pc\n"
            # elif str(sheet.cell(row=x, column=4).value)=="Sundarban's Unprocessed Honey-Half KG":
            #     unit=" pcs\n"
            # elif str(sheet.cell(row=x, column=4).value)=="Kalo Jeera Honey-Half KG":
            #     unit=" pcs\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Sonali Chicken":
            #     unit = " pcs\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Deshi Chicken":
            #     unit = " pcs\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Deshi Morog":
            #     unit = " pcs\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Deshi Duck":
            #     unit = " pcs\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Previous Due":
            #     unit = " \n"
            # elif str(sheet.cell(row=x, column=4).value) == "Transaction Charge":
            #     unit = "- \n"
            # elif str(sheet.cell(row=x, column=4).value) == "Delivery Charge":
            #     unit = "- \n"
            # elif str(sheet.cell(row=x, column=4).value) == "Hilsha Fish-Chadpur":
            #     unit = " pcs\n"
            # elif str(sheet.cell(row=x, column=4).value) == "Hilsha Egg-500gm":
            #     unit = " pcs\n"
            # else:
            #     unit=" KG \n"
            unit = str(sheet.cell(row=x, column=13).value)
            Product = Product + str(sheet.cell(row=x, column=4).value) + "\n"
            Quantity = Quantity + str(sheet.cell(row=x, column=5).value) + " "+unit+"\n"

            TD = TD + sheet.cell(row=x, column=6).value

            Subtotal = Subtotal + str(round(sheet.cell(row=x, column=8).value)) + "/-\n"
            Discount = sheet.cell(row=x, column=12).value
            Total = Total + round(sheet.cell(row=x, column=8).value) - Discount
            if str(sheet.cell(row=x, column=4).value) == "Transaction Charge":
                # Subtotal = " \n"
                Priceperkg = Priceperkg + str(sheet.cell(row=x, column=7).value) + "- \n"
                # Quantity = " \n"
            elif str(sheet.cell(row=x, column=4).value) == "Delivery Charge":
                Priceperkg = Priceperkg + str(sheet.cell(row=x, column=7).value) + "- \n"
            else:
                Priceperkg = Priceperkg + str(sheet.cell(row=x, column=7).value) + "/-\n"

        else:
            # print(f'{Name}\'s-\n{Product}')
            # print(f'sub loop {x}')
            if x == None:
                x = max_col
            if Product == None:
                Product = ""
                Quantity = ""
                Priceperkg = ""
                TD = 0
                Total = 0
            P = Product[0:len(Product) - 1]
            Q = Quantity[0:len(Quantity) - 1]
            Pr = Priceperkg[0:len(Priceperkg) - 1]
            Sl = Subtotal[0:len(Subtotal) - 1]
            imp = [x, P, Q, Pr, Total, TD, Sl]
            return imp


a = True
i = 1
spa = "\n"
while a:
    name = str(sheet.cell(row=i, column=1).value)
    address = str(sheet.cell(row=i, column=3).value)
    contact = str(sheet.cell(row=i, column=2).value)
    Date = sheet.cell(row=i, column=9).value
    paystts = str(sheet.cell(row=i, column=10).value)
    discount_campaign = str(sheet.cell(row=i, column=11).value)
    dsc = sheet.cell(row=i, column=12).value
    Area = sheet.cell(row=i, column=13).value
    discount_amount = "-"+str(dsc) + "/-"


    print(name)
    m = customer_name(name, i)

    i = m[0]
    if i == max_col:
        a = False
    print(f'{m[1]}{m[2]}{m[4]}{m[5]}{spa}')
    tot = f'{m[4]+m[5]}/-'
    d = f'{m[5]}/-'
    document1 = MailMerge(template)
    if paystts != "Online":
        paystts = tot
    if dsc == 0:
        discount_amount=""


    document1.merge(
        Name=name,
        address=address,
        contact=contact,
        product=m[1],
        quantity=m[2],
        pricepp=m[3],
        td=d,
        subtotal= m[6],
        total=tot,
        date=Date.strftime("%d %b, %Y"),
        # order=f'{m[1][0]}{Date.strftime("%d%m%y")}{name[0]}{contact[9:11]}',
        order=f'NIT-{Date.strftime("%d%m")}-{contact[7:11]}',
        COD=paystts,
        discount= f'{discount_campaign}',
        dsc = discount_amount,
        area = Area
    )
    # document1.write(f'invoice\doc\mango\Fahad\{Date.strftime("%b-%d")}{name}.docx')


    document1.write(f'invoice\doc\{Date.strftime("%b-%d")}_{name}.docx')
    # convert(f'invoice\doc\{Date.strftime("%b-%d")}_{name}.docx', f'invoice\pdf\{Date.strftime("%b-%d")}_{name}.pdf')